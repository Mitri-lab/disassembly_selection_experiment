import sys
import string
import numpy as np
import pandas as pd
import selection_funcs_v010 as selfun
import openpyxl

from openpyxl.workbook import Workbook
from openpyxl.styles import PatternFill # Font, Fill,

##### Define experimental parameters
Nspec  = 11  # Number of species (before evolution)
Ntubes = 30  # Number of tubes per experiment condition
Nblank = 1   # Number of blanks (negative controls)
Ntop   = 10  # Number of communities to propagate to next round
scorepow = 1.0
penpow = 1.0
N_spc = 4
pmigr = 0.7

##### TEMPORARY: In pilot, there are 27 tubes, 1 full co-culture and 2 blanks
Ntubepilot = Ntubes -Nblank

##### Define paths for input files
#basedir  = '/Users/bvessman/gitfolder/Selection/data/'
if len(sys.argv)<2:
    print('Input error: You need to specify which input and output files to use. ')
    print('$ python3 get_disassembly_Monday ./path_to_input/input.xlsx')
    quit()

#### Load dataset
collabels = ('D:H,J:M')
#datafile = '20190718_4SC-ASE_Pilot1.xlsx'
#datadict_raw = pd.read_excel(join(basedir,datafile), header=0, sheet_name='Raw', usecols=collabels)

datafile = sys.argv[1]
print('Reading from '+datafile)
datadict_raw = pd.read_excel(datafile, header=0, sheet_name='Raw', usecols=collabels)

#### Split into selection and control
coltup_Wed = ('Tube','Sp_Nr','Sp_Presence','COD_Day0','COD_Day4','Disassembly_Agar','Survival_Day4','Extin-Cont_Day4')

#### Soon, these will be treatment 1 and 2 to blind
df_sel  = datadict_raw.loc[datadict_raw.Selection_Treatment=='Group',coltup_Wed].copy()
df_ctrl = datadict_raw.loc[datadict_raw.Selection_Treatment=='Random',coltup_Wed].copy()

#### Define Survival, extinction etc to column 'Extin-Cont_Day4'
conditions = ('Sp_Presence==1 and Survival_Day4==1',
              'Sp_Presence==1 and Survival_Day4==0',
              'Sp_Presence==0 and Survival_Day4==1',
              'Sp_Presence==0 and Survival_Day4==0')
labeltup   = ('Survival','Exctinction','Contamination','Absent')
for df in (df_sel,df_ctrl):
    for cond, label in zip(conditions,labeltup):
        idces = df.query(cond).index
        for idx in idces:
            df.loc[idx,'Extin-Cont_Day4'] = label

### In selection treatment: replicate communities based on
pardict= {'scorepow':scorepow,'penpow':penpow,'N_spc':N_spc,'Ntubes':Ntubepilot}

#### Proposed communities based on COD scores
dftubes_sel, dfpm_sel_nochange = selfun.selection_treatment_premigr(df_sel,pardict)
dfpm_ctrl_nochange             = selfun.control_treatment_premigr(df_ctrl,pardict)

specpool = tuple(np.unique([str(s) for s in df_sel.Sp_Nr]))

spec_exclude = {'2':'122','122':('2','101'),'3':('1w','108'),'1w':'3',
                '101':'122','103':'114','108':'3','114':'103'}
pardict  = {'specpool':specpool,'N_spc':N_spc,'premix':pmigr,'specexcl':spec_exclude}

#### Finally - copy before-migration communities and migrate!
notpres_sel  = [1,]
while len(notpres_sel)>0:
    df_premigr_sel = dfpm_sel_nochange.copy()
    notpres_sel  = selfun.migrate_members(df_premigr_sel,pardict)

notpres_ctrl = [1,]
while len(notpres_ctrl)>0:
    df_premigr_ctrl = dfpm_ctrl_nochange.copy()
    notpres_ctrl = selfun.migrate_members(df_premigr_ctrl,pardict)

#### Then, define two output formats and write out results to xlsx
coltup = tuple(['Sp_'+str(s) for s in np.arange(1,N_spc+1)])

#### First format: list composition by tube
# outfile = './newcomms_bytube_190807.xlsx'
outfile = sys.argv[2]
print('Writing first format, species by tubes, to '+outfile)
with pd.ExcelWriter(outfile,engine='openpyxl',date_format='YYYYMMDD', mode='w') as writer:
    df_premigr_sel.loc[:,coltup].astype('str').to_excel(writer,sheet_name='Treatment 1')
    df_premigr_ctrl.loc[:,coltup].astype('str').to_excel(writer,sheet_name='Treatment 2')

#### Second format: list tubes by species
occ_sel = selfun.get_tubesperspec(df_premigr_sel)
occ_ctrl = selfun.get_tubesperspec(df_premigr_ctrl)
specorder = ['1','2','122','3','119','1w','101','103','108','109','114']
#specorder = ['1w','101','103','108','109','114','1','2','122','3','119']

#### Define
wb = Workbook()
ws1 = wb.create_sheet("Treatment 1",0)
ws2 = wb.create_sheet("Treatment 2",1)

colors_greyscale = ('FFFFFF','DDDDDD')
colors_treatment = ('E2EFDA','E25EA6')

#### 52 columns should be enough, sorry future me for the unexpected bug.
colarr_Exc = np.append(np.array(list(string.ascii_uppercase)),np.array([c+c for c in string.ascii_uppercase]))

for occurr,ws,offset,scol in zip((occ_sel,occ_ctrl),(ws1,ws2),[0,Ntubes],colors_treatment):
    fillcol_spec = openpyxl.styles.colors.Color(rgb=scol)
    for idxs,spec in enumerate(specorder):
        fillcol_grey = openpyxl.styles.colors.Color(rgb=colors_greyscale[idxs%2])
        rowexc = idxs+1
        ws['A'+str(rowexc)] = spec
        ws['A'+str(rowexc)].fill = PatternFill(patternType='solid', fgColor=fillcol_spec)

        for idxt, tube in enumerate(occurr[spec]):
            colexc = colarr_Exc[idxt+1]
            cell = str(colexc)+str(rowexc)
            ws[cell] = tube
            ws[cell].fill = PatternFill(patternType='solid', fgColor=fillcol_grey)

# outfile = './tubesperspec_Thursday_190808.xlsx'
outfile = sys.argv[3]
print('Writing second format, tubes by species, to '+outfile)
wb.save(outfile)
