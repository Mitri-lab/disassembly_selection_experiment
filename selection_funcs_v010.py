######### Supporting functions for selection experiment algorithm #########
#### By Björn Vessman, Bjorn.Vessman@unil.ch at the Mitri Lab, UNIL DMF
#### Versions
## 0.10    Initial version: basic functionality to read, write data and
##         perform community selection step

#### Standard libraries needed
import os
import string
from os import listdir
from os.path import isfile, join
from datetime import date, datetime

#### Computing libraries
# import matplotlib.pyplot as plt
import scipy
import random
import numpy as np
import pandas as pd
from numpy import linalg as LA
from scipy import stats

#### I/O library for read/write of MS Excel files
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.styles import Font, Fill, PatternFill
from openpyxl import load_workbook

##########################################################################

##### Any global variables
specorder = [1,2,122,3,119,'1w',101,103,108,109,114]

##########################################################################


def idxJaccard(comm1,comm2):
    set1 = set(comm1)
    set2 = set(comm2)
    inter12 = set1.intersection(set2)
    union12 = set1.union(set2)
    #if len(inter12) != len(union12):
    #    print((inter12,union12))
    return(np.divide(len(inter12),len(union12)))

def specdiversity(specpres,q):
    ### At a particular transfer, calculate
    ######    Shannon diversity H' = -sum[ p*log(p) ]
    ######    Gamma diversity   G  =  ( sum[p*p^(q-1)] )^(1/(1-q))

    spec_uniq = np.unique(specpres)
    pi = [np.divide(np.sum(s==specpres),len(specpres)) for s in spec_uniq]

    H = -1.0*np.dot(pi,np.log(pi))
    # G = np.power(np.dot(pi,np.power(pi,q-1)),1.0/(1-q))
    G = np.power(np.sum(np.power(pi,q)),1.0/(1-q))
    return([H,G])


##########################################################################
##########################################################################


def get_veryfirst_tubes(exppars):
    #### Create a set of fresh communities incrementally,
    ##   taking into account that some cannot be combined

    Nspc = exppars['Nspc']
    tubelist = exppars['tubevec']
    specpool = exppars['specpool']
    spec_exclude = exppars['spec_exclude']

    Ntubes = len(tubelist)
    comm  = ['',]*Nspc
    tubes = [comm]*Ntubes
    for idxt, t in enumerate(tubelist):
        s1 = np.random.choice(specpool, size=1)[0]
        comm = [s1,'','','']
        for idxs in [1,2,3]:
            pooltmp = set(specpool)-set(comm)
            for s in comm:
                if s in spec_exclude.keys():
                    incomplist = spec_exclude[s]
                    if type(incomplist) is str:
                        incomplist = [incomplist,]
                    pooltmp -= set(incomplist)
            comm[idxs] = np.random.choice(list(pooltmp), size=1)[0]
        tubes[idxt] = comm

    df_tubes = pd.DataFrame(index=tubelist,data=tubes)
    df_tubes.columns = tuple(['Sp_'+str(i+1) for i in range(Nspc)])
    return(df_tubes)


##########################################################################
##########################################################################


def make_disassembly_dfs_minifile(dfpres,params):
    ###### On Mondays, collect COD data from Day 4 to determine which tubes to disassemble
    ####     The top 10 (33%) of tubes are chosen for selection, as long as all species are present in these comms
    ####     If not all species are present in the chosen comms, reshuffle until all species are.
    #### INPUT
    ## dfpres    pd.DataFrame(index='Tube',columns=('Sp_1','Sp_2','Sp_3','Sp_4','COD_Day4'))
    ##           containing presence of each species in each tube, the associated COD
    ## params    dict of parameters:
    ##             'Nspec'   how many species?
    ##             'Ntop'    how many tubes should be disassembled
    #### OUTPUT
    ## rep       integer: how many plates will we use to disassemble the communities? (For debug purposes)
    ## dfpiv     pd.DataFrame() pivoted from dfpres with Tube numbers and the corresponding
    ##             Species presence/absence in tube
    ##             Marker for disassembly
    def find_da_tubes(specpres_topN, Ntop, Nnontop):
        ## Find tube combination for disassembly, such that:
        ##   -all species are plated, and
        ##   -combination represents as high COD as possible
        #### INPUT
        ## specpres_topN   pd.Series() Sum of species presence in the Ntop communities with highest COD scores
        ## Ntop            Number of communities considered
        ## Nnontop         Ntubes = Nnontop +Ntop

        specpool = set([str(s) for s in specpres_topN.index])
        specbest = set([str(s) for s in specpres_topN.index if specpres_topN[s]>0])
        idxbest  = np.arange(0,Ntop)

        k_backw = 0
        while k_backw<=Ntop and len(specbest)<len(specpool):
            k_backw+=1
            idx10tmp = idxbest.copy()
            for l_forw in np.arange(1,Nnontop,1):
                idx10tmp[Ntop-k_backw] = Ntop+l_forw
                specpres10 = dfpiv.iloc[idx10tmp,:Nspec].sum(axis=0)
                M10 = set([str(s) for s in specpres10.index if specpres10[s]>0])
                if len(M10)>len(specbest):
                    specbest = M10
                    idxbest  = idx10tmp.copy()
        return(idxbest)

    #### Read prelims
    Nblanks = params['Nblanks']
    Ntubes  = params['Ntubes']
    Nspec   = params['Nspec']
    Ntop    = params['Ntop']
    N_spc   = params['N_spc']
    speclist = params['speclist']
    speccols = ['Sp_'+str(r) for r in np.arange(1,N_spc+1)]

    #### Which tubes are there? (Probably 1-30 or 31-60)
    tubelist = np.unique(dfpres.index)
    codseries = dfpres.loc[:,'COD_Day4']
    codvec    = codseries.values

    #### Pivot presence-absence data and add COD column
    dfsort   = dfpres.sort_values(by='COD_Day4')

    dfpiv = pd.DataFrame(data=np.zeros((Ntubes-Nblanks,Nspec)), index=dfsort.index, columns=speclist)
    for t in dfsort.index:
        for s in speccols:
            dfpiv.loc[t,str(dfsort.loc[t,s])] = 1.0

    #### Traversing list by COD scores, find which communities to plate and disassemble
    specpres = dfpiv.iloc[:Ntop,:].sum(axis=0)
    idx_bestcomb = find_da_tubes(specpres, Ntop, len(tubelist)-Ntop)

    #### Add suggestion for which tubes to disassemble
    diassembl_vec = np.zeros(len(tubelist))
    diassembl_vec[idx_bestcomb] = np.ones(Ntop)
    # dfpiv = dfpiv.loc[:,[str(s) for s in specorder]]
    dfpiv.loc[:,'Disassembly_Agar'] = diassembl_vec

    return(dfpiv.sort_index())



##########################################################################
##########################################################################



def make_disassembly_dfs(dfpres,params):
    ###### On Mondays, collect COD data from Day 4 to determine which tubes to disassemble
    ####     The top 10 (33%) of tubes are chosen for selection, as long as all species are present in these comms
    ####     If not all species are present in the chosen comms, reshuffle until all species are.
    #### INPUT
    ## dfpres    pd.DataFrame(columns=('Tube','Sp_Nr','Sp_Presence','COD_Day4'))
    ##           containing presence of each species in each tube, the associated COD
    ## params    dict of parameters:
    ##             'Nspec'      how many species?)
    ##             'maxplates'  how many tubes can be disassembled, at maximum?
    ##             'minreps'    how many replicates of each species, at minimum?
    #### OUTPUT
    ## rep       integer: how many plates will we use to disassemble the communities? (For debug purposes)
    ## dfpiv     pd.DataFrame() pivoted from dfpres with Tube numbers and the corresponding
    ##             Species presence/absence in tube
    ##             COD measurement
    ##             Marker for disassembly
    def find_da_tubes(specpres_topN, Ntop, Nnontop):
        ## Find tube combination for disassembly, such that:
        ##   -all species are plated, and
        ##   -combination represents as high COD as possible
        #### INPUT
        ## specpres_topN   pd.Series() Sum of species presence in the Ntop communities with highest COD scores
        ## Ntop            Number of communities considered
        ## Nnontop         Ntubes = Nnontop +Ntop

        specpool = set([str(s) for s in specpres_topN.index])
        specbest = set([str(s) for s in specpres_topN.index if specpres_topN[s]>0])
        idxbest  = np.arange(0,Ntop)

        k_backw = 0
        while k_backw<=Ntop and len(specbest)<len(specpool):
            k_backw+=1
            idx10tmp = idxbest.copy()
            for l_forw in np.arange(1,Nnontop,1):
                idx10tmp[Ntop-k_backw] = Ntop+l_forw
                specpres10 = dfpiv.iloc[idx10tmp,:Nspec].sum(axis=0)
                M10 = set([str(s) for s in specpres10.index if specpres10[s]>0])
                if len(M10)>len(specbest):
                    specbest = M10
                    idxbest  = idx10tmp.copy()
        return(idxbest)

    #### Read prelims
    Nspec = params['Nspec']
    Ntop  = params['Ntop']

    #### Which tubes are there? (Probably 1-30 or 31-60)
    tubelist = np.unique(dfpres.Tube)
    codvec  = [dfpres.loc[dfpres.Tube==t,'COD_Day4'].values[0] for t in tubelist]
    codseries = pd.Series(index=tubelist,data=codvec)

    #### Pivot presence-absence data and add COD column
    dfpiv = dfpres.pivot(index='Tube', columns='Sp_Nr', values='Sp_Presence').copy()
    # dfpiv.loc[:,'COD_Day4'] = codvec
    # dfpiv = dfpiv.sort_values(by='COD_Day4')
    dfpiv = dfpiv.loc[codseries.sort_values().index,:]

    #### Traversing list by COD scores, find which communities to plate and disassemble
    specpres = dfpiv.iloc[:Ntop,:Nspec].sum(axis=0)
    idx_bestcomb = find_da_tubes(specpres, Ntop, len(tubelist)-Ntop)

    #### Add suggestion for which tubes to disassemble
    diassembl_vec = np.zeros(len(tubelist))
    diassembl_vec[idx_bestcomb] = np.ones(Ntop)
    dfpiv = dfpiv.loc[:,specorder]
    dfpiv.loc[:,'Disassembly_Agar'] = diassembl_vec

    return(dfpiv.sort_index())


##########################################################################
##########################################################################


def cellcolorExcel_disassembly(infname,dfs,colors,idxoffset,colcols=['A',],name_ext=''):
    ###### A function to color the cells of an Excel workbook with disassembly data
    #### INPUT
    ##   infname    The filename (including path!) of the Excelfile to modify.
    ##   name_ext   String of filename extension, if any
    ##   dfs        pd.DataFrames output by the 'function make_disassembly_dfs(dfpres,params)'
    ##   colors     RBG hex values for the background color of the cells in question
    ##   colcols    Which columns to colour
    ##   idxoffset  Offsets of the python indexer (starting at 0) compared to the Excel row number
    ##                The control treatment tubes are numbered 31-60, so needs a modifier -30
    ##     --NOTE-- the dfs, colors and idxoffset are iterables of the same size --NOTE--
    ##
    #### OUTPUT:  None

    wb = load_workbook(infname)
    for color, sheet, dfpiv, idxoff in zip(colors,wb.sheetnames,dfs,idxoffset):
        fillcolor = openpyxl.styles.colors.Color(rgb=color)

        ws = wb[sheet]
        for idx in dfpiv.loc[dfpiv.Disassembly_Agar>0,'Disassembly_Agar'].index:
            # celltmpA = 'A'+str(idx+idxoff)
            # ws[celltmp].fill = PatternFill(patternType='solid', fgColor=fillcolor)
            for letter in colcols:
                celltmp = letter+str(idx+idxoff)
                ws[celltmp].fill = PatternFill(patternType='solid', fgColor=fillcolor)

    dirname  = os.path.dirname(infname)
    basefname,fext = os.path.splitext(os.path.basename(infname))
    outfname = os.path.join(dirname, basefname +name_ext +fext)
    wb.save(outfname)
    return()


##########################################################################
##########################################################################


def readwrite_CFU_COD(df_raw,dfCFUCOD,dfpivs,outfile):
    # specorder = ['1','2','122','3','119','1w','101','103','108','109','114']
    specorder_Excl = ['1w','1','2','3','101','103','108','109','114','119','122']

    #### Input dfs
    dfCFU, dfCOD = dfCFUCOD
    dfpiv_sel, dfpiv_ctrl = dfpivs

    #### CFU, COD are repeated
    CFUvec = dfCFU.set_index('Species').loc[specorder_Excl]

    df_raw.loc[:,'log10CFUml_Day0'] = np.tile(CFUvec.values.transpose()[0],len(dfCOD.index))
    df_raw.loc[:,'COD_Day0'] = np.repeat(dfCOD.loc[:,'COD0'].values,len(specorder_Excl))
    df_raw.loc[:,'COD_Day4'] = np.repeat(dfCOD.loc[:,'COD4'].values,len(specorder_Excl))

    #### Disassembly?
    disasseries = dfpiv_sel.Disassembly_Agar
    disasseries = disasseries.append(dfpiv_ctrl.Disassembly_Agar)
    df_raw.loc[:,'Disassembly_Agar'] = np.repeat(disasseries.values,len(specorder_Excl))

    # print(df_raw.loc[:,('COD_Day0','log10CFUml_Day0')].head(15))
    # outfile = join(basedir,'20190718_4SC-ASE_Pilot1_fakeMonday.xlsx')
    # outfile = datafile
    with pd.ExcelWriter(outfile,engine='openpyxl',date_format='YYYYMMDD', mode='w') as writer:
        df_raw.to_excel(writer,sheet_name='Raw')
    return()


##########################################################################
##########################################################################


def get_spec_in_comm(df_comm):
    #### From a dataframe df_comm with species columns and community labels,
    ##   record where each species feature and return
    ##### INPUT
    ##   df_comm    pandas.DataFrame(index='label',columns=('sp_1','sp_2' ... )) OR
    ##              pandas.DataFrame(columns=('label','sp_1','sp_2' ... ))
    ##### OUTPUT
    ##   comm_dict           dictionary version of df_comm list of species per community
    ##   dict_spec_ocurr     dictionary with presence of species S_i in communities Ck
    ##              {'S_1':{'C1','C2','C4'...}, 'S_2':{'C3','C4','C5'...}, ...}

    commlabels = df_comm.index # ! Note ! some df.s may have the labels as a column
    Ncomms     = len(commlabels)
    columns_sp = [c for c in df_comm.columns if 'sp' in c]

    dict_spec_ocurr = dict()
    comm_dict       = dict()
    for clabel in commlabels:
        comm = df_comm.loc[clabel,columns_sp].values

        comm_dict[clabel] = tuple()
        for s in comm:
            # Convert data type of species label to str
            if type(s) is str:
                slab = s
            elif type(s) is int:
                slab = str(s)
            elif type(s) is np.int64:
                slab = str(s)
            elif type(s) is float:
                slab = str(int(s))
            elif type(s) is np.float64:
                slab = str(int(s))
            else:
                print('Unknown datatype: '+str(type(s)))

            # Record which communities each species featured in
            if slab in dict_spec_ocurr.keys():
                dict_spec_ocurr[slab] += (clabel,)
            else:
                dict_spec_ocurr[slab] = (clabel,)

            # Re-record species presence in communities
            comm_dict[clabel] += (slab,)

    return(comm_dict, dict_spec_ocurr)


##########################################################################
##########################################################################


def find_species_from_tubes_minifile(dfpres, dfsurv, dfCOD):
    ###### On Wednesday, decide which to collect which species from which tube
    ####   Each species is found from its highest-COD tube
    #### INPUT
    ## dfpres     pd.DataFrame(index='Tube',columns=('S1','S2', ...,'SN'))
    ##                containing presence at day 0 of each species
    ## dfsurv     pd.DataFrame(index='Tube',columns=('S1','S2', ...,'SN'))
    ##                containing presence at day 4 of each species in each _plated_ tube
    ## dfCOD      pd.DataFrame(index='Tube',colums=('COD_Day0','COD_Day4'))
    ##                COD measurement
    #### OUTPUT
    ## platedict  pd.Series: From where shall we plate the bacteria?

    #### Which tubes were plated? Which COD for these tubes?
    tubelist = list(set(dfsurv.index).intersection(set(dfpres.index)))
    dfpres_sub = dfpres.loc[tubelist,dfsurv.columns].copy()

    #### Consider only those species that 1) are present, and 2) were supposed to
    dfpiv = dfpres_sub.multiply(dfsurv.loc[tubelist,:]).copy()
    dfpiv.loc[:,'COD'] = dfCOD.loc[tubelist,'COD4'].copy()

    #### Traversing list by COD scores, find from which communities species should be propagated
    platedict   = dict()
    for s in dfsurv.columns:
        platedict[str(s)] = 0
        dfsubset = dfpiv.loc[dfpiv.loc[:,s]>0,'COD']
        if len(dfsubset)>0: # Tmp since not all species were plated in pilot
            platedict[str(s)] = dfsubset.idxmin()

    return(pd.Series(platedict))


def colorExcel_specfromtubes_mf(dfs,idxoffsets,infile,outfile):
    #### Given an MS Excel file ('infile') that was generated by the script
    ##   get_disassembly_Monday(), fill the cells that correspond to a tube,
    ##   from where a certain species should be isolated

    colors = ('E2EFDA','E25EA6')
    colorder_exc = {'Tube':'A','1':'B','2':'C','122':'D','3':'E','119':'F','1w':'G','101':'H','103':'I','108':'J','109':'K','114':'L'}
    # ['1','2','122','3','119','1w','101','103','108','109','114']

    #### Find which Excel cells to use. For now assuming a particular structure.
    excells = list([list()]*2)
    dftreats, dfsurv, dfCOD = dfs
    for idx, (offs, df) in enumerate(zip(idxoffsets,dftreats)):
        # findspecs = find_species_from_tubes(df)
        findspecs = find_species_from_tubes_minifile(df, dfsurv, dfCOD)
        findspecs = findspecs.loc[findspecs>0]
        excelrows = np.add(findspecs,offs) 
        excells[idx] = [colorder_exc[str(i)]+str(r) for i,r in excelrows.items()]

    #### Manipulate Excel workbook
    wb = load_workbook(infile)

    for color, sheet, cellvec in zip(colors,wb.sheetnames,excells):
        fillcolor = openpyxl.styles.colors.Color(rgb=color)
        ws = wb[sheet]
        for cell in cellvec:
            ws[cell].fill = PatternFill(patternType='solid', fgColor=fillcolor)

    wb.save(outfile)
    return()


def find_species_from_tubes(dfpres):
    ###### On Wednesday, decide which to collect which species from which tube
    ####     Each species is found from the tube at which it is present, with the highest COD
    #### INPUT
    ## dfpres       pd.DataFrame(columns=('Tube','Sp_Nr','Sp_Presence','COD_Day4'))
    ##                containing presence of each species in each tube, the associated COD
    ## df_disas     pd.DataFrame() computed from make_disassembly_dfs()
    ##                Species presence/absence in tube
    ##                COD measurement
    ##                Marker for disassembly
    #### OUTPUT
    ## platedict    pd.Series: From where shall we plate the bacteria?

    #### Find pivoted dataframe with presence-absence and pivot
    df_disassem = dfpres.loc[dfpres.query('Disassembly_Agar==1').index,:]

    #### Find who was supposed to be present and who actually was presence-absence data and add COD column
    dfpiv_Pres = df_disassem.pivot(index='Tube', columns='Sp_Nr', values='Sp_Presence').copy()
    dfpiv_Surv = df_disassem.pivot(index='Tube', columns='Sp_Nr', values='Survival_Day4').copy()
    dfpiv = dfpiv_Pres.multiply(dfpiv_Surv).copy()

    #### Which tubes are there? (Probably 1-30 or 31-60)
    tubelist = np.unique(dfpiv.index)
    codvec   = [df_disassem.loc[df_disassem.Tube==t,'COD_Day4'].median() for t in tubelist]
    dfpiv.loc[tubelist,'COD_Day4'] = codvec

    #### Traversing list by COD scores, find from which communities species should be propagated
    platedict   = dict()
    vec_species = dfpres.loc[dfpres.Tube==dfpres.Tube.min(),'Sp_Nr']
    for s in vec_species:
        platedict[s] = 0
        dfsubset = dfpiv.loc[dfpiv.loc[:,s]>0,'COD_Day4']
        if len(dfsubset)>0: # Tmp since not all species were plated in pilot
            platedict[s] = dfsubset.idxmin()

    return(pd.Series(platedict))


def colorExcel_specfromtubes(dfs,idxoffsets,infile,outfile):
    #### Given an MS Excel file ('infile') that was generated by the script
    ##   get_disassembly_Monday(), fill the cells that correspond to a tube,
    ##   from where a certain species should be isolated

    colors = ('E2EFDA','E25EA6')
    colorder_exc = {'1':'B','2':'C','122':'D','3':'E','119':'F','1w':'G','101':'H','103':'I','108':'J','109':'K','114':'L'}
    # ['1','2','122','3','119','1w','101','103','108','109','114']

    #### Find which Excel cells to use. For now assuming a particular structure.
    excells = list([list()]*2)
    for idx, (offs, df) in enumerate(zip(idxoffsets,dfs)):
        findspecs = find_species_from_tubes(df)
        findspecs = findspecs.loc[findspecs>0]
        excelrows = np.add(findspecs,offs)
        excells[idx] = [colorder_exc[str(i)]+str(r) for i,r in excelrows.items()]

    #### Manipulate Excel workbook
    wb = load_workbook(infile)

    for color, sheet, cellvec in zip(colors,wb.sheetnames,excells):
        fillcolor = openpyxl.styles.colors.Color(rgb=color)
        ws = wb[sheet]
        for cell in cellvec:
            ws[cell].fill = PatternFill(patternType='solid', fgColor=fillcolor)

    wb.save(outfile)
    return()


##########################################################################
##########################################################################


def get_master_df(dftup):
    #### Collect a 'master' DataFrame from presence, COD and survival

    (df_pres_T1,df_pres_T2),df_surv,df_COD = dftup

    tubelist_da = df_surv.index
    speclist    = df_surv.columns 
    Ntubes_da = len(tubelist_da)
    Nspec     = len(speclist)

    #### Find dataframe of Treatment 1
    tubelist1 = list(set(df_surv.index).intersection(set(df_pres_T1.index)))
    # Sara changed syntax of all melt() functions below (4x) to pd.melt(..)
    dftmp1 = pd.melt(df_pres_T1.loc[tubelist1,speclist]).copy()
    dftmp1 = dftmp1.rename(columns={'value':'Sp_Presence','variable':'Sp_Nr'})
    dftmp1.loc[:,'Tube']  = np.tile(tubelist1,Nspec)
    dftmp1.loc[:,'Survival_Day4'] = pd.melt(df_surv.loc[tubelist1,:]).value
    # dftmp1 = dftmp1.sort_values(by='Tube')

    #### Find dataframe of Treatment 2
    tubelist2 = list(set(df_surv.index).intersection(set(df_pres_T2.index)))
    dftmp2 = pd.melt(df_pres_T2.loc[tubelist2,speclist]).copy()
    dftmp2 = dftmp2.rename(columns={'value':'Sp_Presence','variable':'Sp_Nr'})
    dftmp2.loc[:,'Tube']  = np.tile(tubelist2,Nspec)
    dftmp2.loc[:,'Survival_Day4'] = pd.melt(df_surv.loc[tubelist2,:]).value
    # dftmp2 = dftmp2.sort_values(by='Tube')

    dfpres_tmp = dftmp1.append(dftmp2,ignore_index=True).sort_values(by=['Tube','Sp_Nr'])
    df_master = pd.DataFrame({'Sp_Nr':np.tile(speclist,Ntubes_da),
                              'COD_Day0':np.repeat(df_COD.loc[tubelist_da,'COD0'],Nspec),
                              'COD_Day4':np.repeat(df_COD.loc[tubelist_da,'COD4'],Nspec),
                              })

    df_master = df_master.sort_values(by=['Tube','Sp_Nr']).reset_index()
    df_master.loc[:,'Sp_Presence'] = dfpres_tmp.reset_index().loc[:,'Sp_Presence']
    df_master.loc[:,'Survival_Day4'] = dfpres_tmp.reset_index().loc[:,'Survival_Day4']

    return(df_master)


##########################################################################
##########################################################################


def selection_treatment_premigr(df_indata,params):
    ##### Propose a new species composition for tube 1, 2, 3 ..., Ntubes
    ##### In selection treatment, communities are replicated at random, roughly in proportion to COD scores
    ##### from last week. NOTE! This step is followed by migration of randomly chosen species.
    #### INPUT
    ## df_indata    Raw-ish data file with all info on species presence, COD0, COD4 etc.
    ## params       Experimental parameters

    #### Read experimental parameters
    spow  = params['scorepow']
    ppow  = params['penpow']
    N_spc = params['N_spc']

    #### Define new pd.DataFrame()
    dftubes = pd.DataFrame(index=np.unique(df_indata.loc[df_indata.Disassembly_Agar>0,:].Tube))
    Ntubes = len(np.unique(df_indata.Tube))

    #### For each new tube, append COD data and compute COD scores
    prevspecs_dict = dict()
    for tube in dftubes.index:
        ### COD scores
        COD0 = df_indata.loc[df_indata.Tube==tube,'COD_Day0'].median()
        COD4 = df_indata.loc[df_indata.Tube==tube,'COD_Day4'].median()
        relCOD = np.divide(COD0-COD4,COD0)

        ### Survival
        presence  = df_indata.loc[df_indata.Tube==tube,'Sp_Presence']
        survival  = df_indata.loc[df_indata.Tube==tube,'Survival_Day4']
        survratio = np.divide(survival[presence>0].sum(),presence.sum())

        dftubes.loc[tube,'COD_Day0'] = COD0
        dftubes.loc[tube,'COD_Day4'] = COD4
        dftubes.loc[tube,'CODscore'] = relCOD
        dftubes.loc[tube,'Survival'] = survratio

        ### Previous composition
        dfpres = df_indata.loc[df_indata.Tube==tube,('Sp_Nr','Sp_Presence')]
        specvec = dfpres.loc[dfpres.Sp_Presence>0,'Sp_Nr']
        prevspecs_dict[tube] = tuple(specvec)
        for i,s in enumerate(specvec):
            if (type(s) is float or type(s) is np.float64):
                s = int(s)
            dftubes.loc[tube,'Sp_'+str(i+1)] = str(s)

    #### Compute community score and probability of replication
    dftubes.loc[:,'Score'] = np.multiply(np.power(dftubes.loc[:,'CODscore'],spow),
                                         np.power(dftubes.loc[:,'Survival'],ppow))
    dftubes.loc[:,'Prob']  = np.divide(dftubes.loc[:,'Score'],dftubes.loc[:,'Score'].sum())

    #### Based on computed probabilities, draw new communities and add to new pd.DataFrame
    newtubes  = np.random.choice(list(dftubes.index),Ntubes,replace=True,p=dftubes.Prob)

    #### Define output dataframe with proposed species composition in each community
    dftubes_premigr = pd.DataFrame(index=np.unique(df_indata.Tube),data={'From_tube':newtubes})
    for fromtube, totube in zip(newtubes,dftubes_premigr.index):
        # specvec = df_indata.loc[df_indata.Sp_Presence>0,'Sp_Nr']
        prevspecs = prevspecs_dict[fromtube]
        for i,spec in enumerate(prevspecs):
            if (type(spec) is float or type(spec) is np.float64):
                spec = int(spec)
            dftubes_premigr.loc[totube,'Sp_'+str(i+1)] = str(spec)

    return(dftubes,dftubes_premigr)



def selection_treatment_premigr_mf(df_indata,params):
    ##### Propose a new species composition for tube 1, 2, 3 ..., Ntubes
    ##### In selection treatment, communities are replicated at random, roughly in proportion to COD scores
    ##### from last week. NOTE! This step is followed by migration of randomly chosen species.
    #### INPUT
    ## df_indata    Raw-ish data file with all info on species presence, COD0, COD4 etc.
    ## params       Experimental parameters

    #### Read experimental parameters
    spow  = params['scorepow']
    ppow  = params['penpow']
    N_spc   = params['N_spc']
    tubevec = params['tubevec']

    #### Define new pd.DataFrame()
    dftubes = pd.DataFrame(index=np.unique(df_indata.Tube))
    Ntubes  = len(tubevec)

    #### For each new tube, append COD data and compute COD scores
    prevspecs_dict = dict()
    for tube in df_indata.Tube:
        ### COD scores
        COD0 = df_indata.loc[df_indata.Tube==tube,'COD_Day0'].median()
        COD4 = df_indata.loc[df_indata.Tube==tube,'COD_Day4'].median()
        relCOD = np.divide(COD0-COD4,COD0)

        ### Survival
        presence  = df_indata.loc[df_indata.Tube==tube,'Sp_Presence']
        survival  = df_indata.loc[df_indata.Tube==tube,'Survival_Day4']
        survratio = np.divide(survival[presence>0].sum(),presence.sum())

        dftubes.loc[tube,'COD_Day0'] = COD0
        dftubes.loc[tube,'COD_Day4'] = COD4
        dftubes.loc[tube,'CODscore'] = relCOD
        dftubes.loc[tube,'Survival'] = survratio

        ### Previous composition
        dfpres  = df_indata.loc[df_indata.Tube==tube,('Sp_Nr','Sp_Presence')]
        specvec = dfpres.loc[dfpres.Sp_Presence>0,'Sp_Nr']
        prevspecs_dict[tube] = tuple(specvec)
        for i,s in enumerate(specvec):
            if (type(s) is float or type(s) is np.float64):
                s = int(s)
            dftubes.loc[tube,'Sp_'+str(i+1)] = str(s)

    #### Compute community score and probability of replication
    dftubes.loc[:,'Score'] = np.multiply(np.power(dftubes.loc[:,'CODscore'],spow),
                                         np.power(dftubes.loc[:,'Survival'],ppow))
    dftubes.loc[:,'Prob']  = np.divide(dftubes.loc[:,'Score'],dftubes.loc[:,'Score'].sum())

    #### Based on computed probabilities, draw new communities and add to new pd.DataFrame
    newtubes  = np.random.choice(list(dftubes.index),Ntubes,replace=True,p=dftubes.Prob)

    #### Define output dataframe with proposed species composition in each community
    dftubes_premigr = pd.DataFrame(index=tubevec,data={'From_tube':newtubes})
    for fromtube, totube in zip(newtubes,tubevec):
        # specvec = df_indata.loc[df_indata.Sp_Presence>0,'Sp_Nr']
        prevspecs = prevspecs_dict[fromtube]
        for i,spec in enumerate(prevspecs):
            if (type(spec) is float or type(spec) is np.float64):
                spec = int(spec)
            dftubes_premigr.loc[totube,'Sp_'+str(i+1)] = str(spec)

    return(dftubes,dftubes_premigr)


##########################################################################
##########################################################################


def control_treatment_premigr(df_indata,params):
    ##### Propose a new species composition for tube 1, 2, 3 ..., Ntubes
    #### In control treatment, communities are replicated at random
    #### NOTE! This step is followed by migration of randomly chosen species.

    #### Read params, define a DataFrame of species combinations
    tubestot = np.unique(df_indata.Tube)
    tubevec  = np.unique(df_indata.loc[df_indata.Disassembly_Agar>0,:].Tube)
    Ntubes   = len(tubestot)
    newtubes = np.random.choice(tubevec,Ntubes,replace=True)
    dftubes_premigr = pd.DataFrame(index=np.unique(df_indata.Tube),data={'From_tube':newtubes})

    #### Find species composition of previous tubes
    prevspecs_dict = dict()
    for tube in tubevec:
        dfpres  = df_indata.loc[df_indata.Tube==tube,('Sp_Nr','Sp_Presence')]
        specvec = dfpres.loc[dfpres.Sp_Presence>0,'Sp_Nr']
        prevspecs_dict[tube] = tuple(specvec)

    #### ... and assign communities to new tubes
    for fromtube, totube in zip(newtubes,dftubes_premigr.index):
        prevspecs = prevspecs_dict[fromtube]
        for i,spec in enumerate(prevspecs):
            if (type(spec) is float or type(spec) is np.float64):
                spec = int(spec) # Species labels may be interpreted as floats when read from Excel
            dftubes_premigr.loc[totube,'Sp_'+str(i+1)] = str(spec)

    return(dftubes_premigr)


def control_treatment_premigr_mf(df_indata,params):
    ##### Propose a new species composition for tube 1, 2, 3 ..., Ntubes
    #### In control treatment, communities are replicated at random
    #### NOTE! This step is followed by migration of randomly chosen species.

    #### Read params, define a DataFrame of species combinations
    tubelist_new = params['tubevec']
    Ntubes = len(tubelist_new)
    tubelist_old = np.unique(df_indata.Tube)
    newtubes     = np.random.choice(tubelist_old,Ntubes,replace=True)
    dftubes_premigr = pd.DataFrame(index=tubelist_new,data={'From_tube':newtubes})

    #### Find species composition of previous tubes
    prevspecs_dict = dict()
    for tube in tubelist_old:
        dfpres  = df_indata.loc[df_indata.Tube==tube,('Sp_Nr','Sp_Presence')]
        specvec = dfpres.loc[dfpres.Sp_Presence>0,'Sp_Nr']
        prevspecs_dict[tube] = tuple(specvec)

    #### ... and assign communities to new tubes
    for fromtube, totube in zip(newtubes,dftubes_premigr.index):
        prevspecs = prevspecs_dict[fromtube]
        for i,spec in enumerate(prevspecs):
            if (type(spec) is float or type(spec) is np.float64):
                spec = int(spec) # Species labels may be interpreted as floats when read from Excel
            dftubes_premigr.loc[totube,'Sp_'+str(i+1)] = str(spec)

    return(dftubes_premigr)


##########################################################################
##########################################################################


def migrate_members(df_premigr,params):
    ###### Function to re-mix a certain proportion of the community collection in 'df_premigr'
    #### INPUT
    ## df_premigr    pd.DataFrame(index='Tube',columns=('From_tube','Sp_1','Sp_2','Sp_3','Sp_4'))
    ##               Contains next week's tube number as index and
    ##               tube number of previous tubes, as well as community composition as columns
    ## params        Dictionary of community params

    #### Prelims: community and experimental params
    N_spc    = params['N_spc']
    specpool = params['specpool']
    spec_exclude = params['specexcl']
    premix   = params['premix']

    Ntubes = len(df_premigr.index)

    #### Find out who is present and not over all communities
    collabel_members = tuple(['Sp_'+str(s) for s in np.arange(1,N_spc+1)])
    specpres = np.unique([str(s) for s in df_premigr.loc[:,collabel_members].values.flatten()])
    spec_notpres = tuple(set(specpool)-set(specpres))

    #### Randomly designate communities for migration/remix
    Nremix  = round(premix*Ntubes)
    tubevec_migr = np.random.choice(df_premigr.index, replace=False, size=Nremix)
    # tubevec_migr = np.sort(np.random.choice(range(Ntubes), replace=False, size=Nremix))
    # df_premigr.loc[:,'Migrate'] = np.random.binomial(n=1, p=premix, size=Ntubes)
    df_migrtmp = df_premigr.loc[tubevec_migr,collabel_members]

    #### Remix chosen communities
    for tube in np.random.permutation(tubevec_migr):
        speclist  = df_migrtmp.loc[tube,:].values

        #### Is there a unique presence that we should avoid to emigrate?
        tubes_remain = list(set(tubevec_migr)-set([tube,]))
        specremain = np.unique(df_migrtmp.loc[tubes_remain,:].astype(str).values)
        uniquespec = list(set(speclist)-set(specremain))
        if len(uniquespec)>0: # I.e. if some community contains a unique species - don't remove it
            idxuniqe = [idx for idx,us in enumerate(speclist) if us in uniquespec]
            idxtot = range(len(speclist))
            lsnonunique = list(set(idxtot)-set(idxuniqe))
            idx_emig = np.random.choice(lsnonunique,1)[0]
        else:                 # Otherwise it's fine - emigrate whomever
            idx_emig = np.random.choice(range(len(speclist)),1)[0]
        spec_emig = speclist[idx_emig]
        spec_remain = [s for s in speclist if s != spec_emig]

        #### Define list of immigrants by two criteria:
        ## 1) Do we need to immigrate someone who is not yet present?
        ## 2) Are there incompatibilities, that certain species cannot be combined with
        spec_incomp = [spec_exclude[str(s)] for s in spec_remain if str(s) in spec_exclude.keys()]
        if len(spec_incomp)>0 and type(spec_incomp[0]) is tuple:
            pool_immig = list(set(spec_notpres)-set(spec_incomp[0]))
        else:
            pool_immig = list(set(spec_notpres)-set(spec_incomp))

        if len(pool_immig)>0: # Then 1) there are some not-presents and
                              #      2) not all are incompatible
            spec_immig = np.random.choice(pool_immig,1)[0]
        else: # Else: no not-presents or all not-present are compatible
            if len(spec_incomp)>0 and type(spec_incomp[0]) is tuple:
                pool_immig = list(set(specpool)-set(speclist)-set(spec_incomp[0]))
            else:
                pool_immig = list(set(specpool)-set(speclist)-set(spec_incomp))
            spec_immig = np.random.choice(pool_immig,1)[0]
        speclist[idx_emig] = str(spec_immig)
        df_premigr.loc[tube,collabel_members] = speclist

        #### Re-check species presence
        specpres = np.unique([str(s) for s in df_premigr.loc[:,collabel_members].values.flatten()])
        spec_notpres = tuple(set(specpool)-set(specpres))
    return(spec_notpres)


def migrate_members_new(df_premigr,params):
    #### Prelims: community and experimental params
    N_spc    = params['N_spc']
    specpool = params['specpool']
    spec_exclude = params['specexcl']
    premix   = params['premix']

    Ntubes = len(df_premigr.index)

    #### Find out who is present and not over all communities
    collabel_members = tuple(['Sp_'+str(s) for s in np.arange(1,N_spc+1)])
    specpres = np.unique([str(s) for s in df_premigr.loc[:,collabel_members].values.flatten()])
    spec_notpres = tuple(set(specpool)-set(specpres))

    #### Randomly designate communities for migration/remix
    Nremix  = round(premix*Ntubes)
    tubevec_migr = np.random.choice(df_premigr.index, replace=False, size=Nremix)
    df_migrtmp = df_premigr.loc[tubevec_migr,collabel_members]

    #### Remix chosen communities
    for tube in np.random.permutation(tubevec_migr):
        speclist  = df_migrtmp.loc[tube,:].values

        #### Is there a unique presence that we should avoid to emigrate?
        tubes_remain = list(set(tubevec_migr)-set([tube,]))
        specremain = np.unique(df_migrtmp.loc[tubes_remain,:].astype(str).values)
        uniquespec = list(set(speclist)-set(specremain))
        if len(uniquespec)>0: # I.e. if some community contains a unique species - don't remove it
            idxuniqe = [idx for idx,us in enumerate(speclist) if us in uniquespec]
            idxtot = range(len(speclist))
            lsnonunique = list(set(idxtot)-set(idxuniqe))
            idx_emig = np.random.choice(lsnonunique,1)[0]
        else:                 # Otherwise it's fine - emigrate whomever
            idx_emig = np.random.choice(range(len(speclist)),1)[0]
        spec_emig = speclist[idx_emig]
        spec_remain = [s for s in speclist if s != spec_emig]

        #### Define list of immigrants by two criteria:
        ## 1) Do we need to immigrate someone who is not yet present?
        ## 2) Are there incompatibilities, that certain species cannot be combined with
        spec_incomp = [spec_exclude[str(s)] for s in spec_remain if str(s) in spec_exclude.keys()]
        incomp_flat = []
        for sublist in spec_incomp:
            if type(sublist) is tuple:
                for item in sublist:
                    incomp_flat.append(item)
            elif type(sublist) is str:
                incomp_flat.append(sublist)
            else:
                print('Halp! Uncrecognized format for species.')

        pool_immig = list(set(spec_notpres)-set(incomp_flat))
        if len(pool_immig)>0: # Then 1) there are some not-presents and
                              #      2) not all are incompatible
            spec_immig = np.random.choice(pool_immig,1)[0]
        else:
            pool_immig = list(set(specpool)-set(speclist)-set(incomp_flat))
            spec_immig = np.random.choice(pool_immig,1)[0]
        speclist[idx_emig] = str(spec_immig)
        df_premigr.loc[tube,collabel_members] = speclist

        #### Re-check species presence
        specpres = np.unique([str(s) for s in df_premigr.loc[:,collabel_members].values.flatten()])
        spec_notpres = tuple(set(specpool)-set(specpres))
    return(spec_notpres)


##########################################################################
##########################################################################


def get_tubesperspec(df_premigr):

    colvec = ('Sp_1','Sp_2','Sp_3','Sp_4')
    specpool = np.unique([str(v) for v in df_premigr.loc[:,colvec].values.flatten()])
    occurrence = dict()
    for s in specpool:
        idxarr = df_premigr.index[df_premigr.loc[:,colvec[0]].astype(str)==s]
        for column in colvec[1:]:
            idxarr = idxarr.append(df_premigr.index[df_premigr.loc[:,column].astype(str)==s])
        occurrence[s] = list(idxarr.sort_values())

    return(occurrence)


##########################################################################
##########################################################################


def set_tubesbyspec(dfs,outfile):
    def as_text(element):
        if element is None:
            return("")
        return str(element)

    #### Define constants
    specorder = ['1','2','122','3','119','1w','101','103','108','109','114']
    colors_greyscale = ('FFFFFF','DDDDDD')
    colors_treatment = ('E2EFDA','E25EA6')

    #### 52 columns should be enough, sorry future me for the unexpected bug.
    colarr_Exc = np.append(np.array(list(string.ascii_uppercase)),np.array([c+c for c in string.ascii_uppercase]))

    #### Read input dataframes, the ones to print
    dfmigr_sel, dfmigr_ctrl = dfs
    Ntubes = len(dfmigr_sel.index)

    #### Find species occurrence
    occ_sel  = get_tubesperspec(dfmigr_sel)
    occ_ctrl = get_tubesperspec(dfmigr_ctrl)

    #### Define output Excel workbook
    wb = Workbook()
    ws1 = wb.create_sheet("Treatment 1",0)
    ws2 = wb.create_sheet("Treatment 2",1)

    #### For each occurrence of a species in a tube, note this into the workbook
    for occurr,ws,offset,scol in zip((occ_sel,occ_ctrl),(ws1,ws2),[0,Ntubes],colors_treatment):
        fillcol_spec = openpyxl.styles.colors.Color(rgb=scol)
        for idxs,spec in enumerate(specorder):
            fillcol_grey = openpyxl.styles.colors.Color(rgb=colors_greyscale[idxs%2])
            rowexc = idxs+1
            ws['A'+str(rowexc)] = spec
            ws['A'+str(rowexc)].fill = PatternFill(patternType='solid', fgColor=fillcol_spec)

            #### In this structure, write every tube that a species occur in
            for idxt, tube in enumerate(occurr[spec]):
                colexc = colarr_Exc[idxt+1]
                cell   = str(colexc)+str(rowexc)
                ws[cell] = tube
                ws[cell].fill = PatternFill(patternType='solid', fgColor=fillcol_grey)

        #### Adjust cell widths in column
        #for column_cells in ws.columns:
        #    length = max(len(as_text(cell.value)) for cell in column_cells)
        #    ws.column_dimensions[column_cells[0].column].width = length

    wb.save(outfile)
    return()


##########################################################################


def allocatecomms(pvec,numdict):
    #### FUNCTION TO DRAW Ncom communities with N_spc species per community
    #### INPUT
    ## pvec        vector of probabilities corresponding to a list of species
    ## numdict     dict {'Ncom','N_spc'}
    ##
    #### OUTPUT
    ## commvec     np.ndarray(shape=(Ncom,N_spc)) of communities

    def drawbyindex(specarray):
        subcomm = np.ndarray(shape=(len(specarray),N_spc))
        for i,s in enumerate(specarray):
            svectmp = np.append(speclist[:s],speclist[(s+1):]) # ... of species
            qvectmp = np.append(pvec[:s],pvec[(s+1):])         # ... of probs

            q_ex  = np.divide(qvectmp,sum(qvectmp)) # Probability of other members
            comm3 = np.random.choice(svectmp, size=(N_spc-1), replace=False, p=q_ex)
            subcomm[i] = np.append(comm3,s)
        return(subcomm)

    Nspec = len(pvec)
    Ncom  = numdict['Ncom']
    N_spc = numdict['N_spc']
    commvec  = np.ndarray(shape=(Ncom,N_spc))
    speclist = range(Nspec)

    #### Draw communities
    #### 1) A few communities with guaranteed presence
    idxhalf = min(Ncom,Nspec)//2
    spec_lowprob = np.argpartition(pvec, range(idxhalf))

    subcomm = drawbyindex(spec_lowprob[:idxhalf])
    commvec[:idxhalf,:] = subcomm

    #### 2) Re-draw those that are not present
    spec_present = np.unique(commvec[:idxhalf,:])
    spec_notpres = np.array(list(set(speclist)-set(spec_present)))

    len_redraw = len(spec_notpres)
    if len_redraw>0:
        idxrange = np.arange(idxhalf,idxhalf+len_redraw,1)
        commvec[idxrange,:] = drawbyindex(spec_notpres)

    #### 3) Draw final places
    drawrange = np.arange(start=idxhalf+len_redraw,stop=Ncom,step=1)
    for i in drawrange:    # Draw rest of elements
        commvec[i] = np.random.choice(Nspec, size=N_spc, replace=False, p=pvec)

    return( commvec )

##########################################################################
##########################################################################


def simulcomms_hash(commvec, df_combs):
    ##### Simulate the communities defined in commvec
    #### INPUT
    ## commvec   Array [Ncomm, N_spc] of communities
    #### OUTPUT
    ## commprod  Production, i.e. degradation pontential, of communities

    Ncomm, N_spc = commvec.shape
    scol = [c for c in df_combs.columns if 'Spec' in c]

    commprod = [0.0]*Ncomm
    commidcs = [0.0]*Ncomm
    survival = [0.0]*Ncomm

    for ic, com in enumerate(commvec):
        # print(com)
        comsort = np.sort(com)

        df_subset = df_combs.copy()
        for idx,specs in enumerate(zip(comsort,scol)):
            scom, sdf = specs
            df_subset = df_subset.loc[df_subset.loc[:,sdf]==scom,tuple(scol[(idx+1):])+('Prod','Survival')]

        idxprodsurv = df_subset.reset_index().values[0]

        ### Collect index of community
        commidcs[ic] = idxprodsurv[0]

        ### Collect net productivity of community
        commprod[ic] = idxprodsurv[1]

        ## Collect penalty for extinction
        survival[ic] = idxprodsurv[2]

    return( commprod, survival, commidcs )


##########################################################################
##########################################################################


def probfromranks(spec_dict, commranks, pold, w_old):
    ###### Compute weighted prob from new set of ranks
    #### INPUT
    ## spec_dict   dict {['S_i': (Ck for k s.th. S_i in Ck), for all species S_i]}
    ## commranks   dict {['Ck': rank_of_comm_k, for all communities Ck]}
    ## pold        pd.Series() of previous probs, length: Nspec
    ## w_old       Proportional weight of previous probability
    #### OUTPUT
    ## probnew     pd.Series()

    ##### Compute species scores from community ranks
    rank_av  = np.mean(commranks)
    rank_max = np.max(commranks)
    spec_scores = dict()
    for (spec,comms) in spec_dict.items():
        if len(comms) >0:
            # NOTE: Communities are ranked in increasing order, with best comm as index 0
            #       To compute species probs, we want to reverse and
            #       assign higher score to 'good' comms
            spec_scores[spec] = np.sum([rank_max-commranks[c] for c in comms])/len(comms)
        else:
            spec_scores[spec] = rank_av

    ##### Compute new probability from species scores
    spec_prob = dict()
    scoresum  = sum(spec_scores.values())
    for (spec,rank) in spec_scores.items():
        spec_prob[spec] = np.divide(rank,scoresum)

    ##### Merge with new and old probabilities
    probnew = pold.copy()
    for spec,pn in spec_prob.items():
        probnew.loc[spec] += np.divide(1.0-w_old,w_old)*pn
    probnew *= w_old
    return( pd.Series(probnew) )


##########################################################################
##########################################################################


def rankprobcomm(ranks,scalefac):
    #### Per-community probability of re-mixing, based on community ranks
    #### INPUT
    ## rankseries   pandas.Series() of degradation ranks of communities
    ## scalefac     scaling factor for sigmoid function scipy.specia.expit()
    #### OUTPUT
    ## commprob     pandas.Series() of per-community probability of remixing

    scaledranks = ranks.subtract(len(ranks)/2.0).multiply(scalefac)
    commprob = scipy.special.expit(scaledranks)
    return( commprob )


def probcomm_score(scores,scalefac):
    #### Per-community probability of re-mixing, based on community ranks
    #### INPUT
    ## scoreseries  pandas.Series() of community production as COD value
    ## scalefac     scaling factor for sigmoid function scipy.specia.expit()
    #### OUTPUT
    ## commprob     pandas.Series() of per-community probability of remixing

    scaledscore = scores.subtract(scores.median()).multiply(scalefac)
    commprob    = scipy.special.expit(scaledscore)
    return( commprob )


##########################################################################
##########################################################################


def remix_ctrl(comm_dict):
    ##### Communities are randomly re-mixed
    ####  INPUT
    ##  df_comms  Dictionary of which species feature in which community

    #### Basic idea: For each community
    #### 1) If u < pcomm, remix community
    #### 2) Remove one or more species, replace from pool

    Ncomm = len(comm_dict)
    Nspec = len(pvecstr)
    u1vec = np.random.uniform(low=0.0, high=1.0, size=Ncomm)
    probcomm = pd.Series(data=np.random.choice(Ncomm, size=Ncomm, replace=False), index=comm_dict.keys())

    for (comm,cspec),u1,pc in zip(comm_dict.items(),u1vec,probcomm):
        if u1<=pc:
            #### Control design: emigrate randomly
            emigrant = np.random.choice(cspec, size=1)

            # The species that do not feature in the communities can migrate
            specvec = list(set(pvecstr.index)-set(cspec))

            #### Standard design immigrate uniformly
            immigrant = np.random.choice(specvec, size=1)

            #### Append immigrant and return new community composition
            specvec_out = [s for s in cspec if s != emigrant]
            specvec_out.append(immigrant[0])
            comm_dict[comm] = tuple(specvec_out.copy())
    return()


##########################################################################
##########################################################################


def remix_migrateworstunif(comm_dict,probcomm,pvecstr):
    ##### Low-rank communities re-mixed based on their ranks
    ####  A high rank leads to a low probability of remixing
    ####  INPUT
    ##  df_comms  Dictionary of which species feature in which community
    ##  probcomm  pandas.Series() of per-community probability of remixing
    ##  pvecstr   pandas.Series() of per-species probability of community assignment

    #### Basic idea: For each community
    #### 1) If u < pcomm, remix community
    #### 2) Remove one or more species, replace from pool

    Ncomm = len(comm_dict)
    Nspec = len(pvecstr)
    u1vec = np.random.uniform(low=0.0, high=1.0, size=Ncomm)

    for (comm,cspec),u1,pc in zip(comm_dict.items(),u1vec,probcomm):
        if u1<=pc:
            #### Standard design: emigrate worst-scoring
            strprob_comm = pvecstr[list(cspec)]
            minprob = strprob_comm.loc[strprob_comm==strprob_comm.min()]

            # The species that do not feature in the communities can migrate
            specvec = list(set(pvecstr.index)-set(cspec))

            #### Standard: Immigrate max-probability
            psubset   = pvecstr[specvec]
            # immigrant = psubset.loc[psubset==psubset.max()].index[0]

            #### Alternative designs: immigrate uniformly or by prob
            immigrant = np.random.choice(specvec, size=1)
            # immigrant = np.random.choice(specvec, size=1, p=np.divide(psubset,psubset.sum()) )

            #### Append immigrant and return new community composition
            specvec_out = [s for s in cspec if s != minprob.index[0]]
            specvec_out.append(immigrant)
            comm_dict[comm] = tuple(specvec_out.copy())
    return()

def remix_migrateworstpcomm(comm_dict,probcomm,pvecstr):
    ##### Low-rank communities re-mixed based on their ranks
    ####  A high rank leads to a low probability of remixing
    ####  INPUT
    ##  df_comms  Dictionary of which species feature in which community
    ##  probcomm  pandas.Series() of per-community probability of remixing
    ##  pvecstr   pandas.Series() of per-species probability of community assignment

    #### Basic idea: For each community
    #### 1) If u < pcomm, remix community
    #### 2) Remove one or more species, replace from pool

    Ncomm = len(comm_dict)
    Nspec = len(pvecstr)
    u1vec = np.random.uniform(low=0.0, high=1.0, size=Ncomm)

    for (comm,cspec),u1,pc in zip(comm_dict.items(),u1vec,probcomm):
        if u1<=pc:
            #### Standard design: emigrate worst-scoring
            strprob_comm = pvecstr[list(cspec)]
            minprob = strprob_comm.loc[strprob_comm==strprob_comm.min()]

            # The species that do not feature in the communities can migrate
            specvec = list(set(pvecstr.index)-set(cspec))

            #### Standard: Immigrate max-probability
            psubset   = pvecstr[specvec]
            # immigrant = psubset.loc[psubset==psubset.max()].index[0]

            #### Alternative designs: immigrate uniformly or by prob
            # immigrant = np.random.choice(specvec, size=1)
            immigrant = np.random.choice(specvec, size=1, p=np.divide(psubset,psubset.sum()) )

            #### Append immigrant and return new community composition
            specvec_out = [s for s in cspec if s != minprob.index[0]]
            specvec_out.append(immigrant)
            comm_dict[comm] = tuple(specvec_out.copy())
    return()

def remix_migrateworstbest(comm_dict,probcomm,pvecstr):
    ##### Low-rank communities re-mixed based on their ranks
    ####  A high rank leads to a low probability of remixing
    ####  INPUT
    ##  df_comms  Dictionary of which species feature in which community
    ##  probcomm  pandas.Series() of per-community probability of remixing
    ##  pvecstr   pandas.Series() of per-species probability of community assignment

    #### Basic idea: For each community
    #### 1) If u < pcomm, remix community
    #### 2) Remove one or more species, replace from pool

    Ncomm = len(comm_dict)
    Nspec = len(pvecstr)
    u1vec = np.random.uniform(low=0.0, high=1.0, size=Ncomm)

    for (comm,cspec),u1,pc in zip(comm_dict.items(),u1vec,probcomm):
        if u1<=pc:
            #### Standard design: emigrate worst-scoring
            strprob_comm = pvecstr[list(cspec)]
            minprob = strprob_comm.loc[strprob_comm==strprob_comm.min()]

            # The species that do not feature in the communities can migrate
            specvec = list(set(pvecstr.index)-set(cspec))

            #### Standard: Immigrate max-probability
            psubset   = pvecstr[specvec]
            immigrant = psubset.loc[psubset==psubset.max()].index[0]

            #### Alternative designs: immigrate uniformly or by prob
            # immigrant = np.random.choice(specvec, size=1)
            # immigrant = np.random.choice(specvec, size=1, p=np.divide(psubset,psubset.sum()) )

            #### Append immigrant and return new community composition
            specvec_out = [s for s in cspec if s != minprob.index[0]]
            specvec_out.append(immigrant)
            comm_dict[comm] = tuple(specvec_out.copy())
    return()


##########################################################################
##########################################################################


def repsim_combins(df_combs,exppars):
    ########## FUNCTION TO REPEAT SIMULATION OF COMMUNITIES ##########
    ####### INPUT  #######
    #### df_combs
    #### exppars
    ####### OUTPUT #######
    ##########

    speccols = [c for c in df_combs.columns if 'Spec' in c]
    specvec  = np.unique(df_combs.loc[:,speccols])

    Nspec = len(specvec)
    Nrep  = exppars['Nrep']
    Ncomm = exppars['Ncomm']
    w_old = exppars['w_old']
    N_spc = int(np.max(df_combs.Survival)) # Assuming that at least one combination is coexistent

    pvecstr = [np.divide(1.0,Nspec),]*Nspec
    pveccom = [np.divide(1.0,Ncomm),]*Ncomm

    #### Throughout simulation, we collect per-species and community properties
    pcollstr = [pvecstr]*(Nrep+1)                 # Probability-COLLection of STRains
    pcollcom = [pveccom]*(Nrep+1)                 # Probability-COLLection of COMmunities
    commcoll = [np.zeros([Ncomm,N_spc])]*(Nrep+1) # COMMunity COLLection, ie strain identities in communities
    prodcoll = np.zeros([Nrep,Ncomm])             # PRODuction COLLection, ie per-community production

    #### 1) Allocate Ncomm communities with N_spc species per comm
    ##     randomly according to pvec
    commvec = allocatecomms(pvecstr,{'N_spc':N_spc,'Ncom':Ncomm})
    commcoll[0] = commvec.copy()
    for k in range(Nrep):
        #### 2) Simulate growth and hence community production
        commprod, survival, commidcs = simulcomms_hash(commvec, df_combs)
        prodcoll[k,:] = np.abs(commprod)

        #### 3) Find ranks with respect to community degradation
        netpen = np.divide(survival, N_spc)

        # ranks = np.argsort(np.multiply(commprod,np.max(netpen)))
        # ranks = np.argsort(np.multiply(commprod,np.min(netpen)))
        ranks  = np.argsort(np.multiply(commprod,np.power(netpen,4.0)))
        # ranks = np.argsort(np.multiply(netpen,commprod))

        #### 4) Compute new probabilities
        pvecstr = probfromranks(commvec, ranks, pvecstr, [w_old,1.0-w_old])
        pveccom = probcomm_score(commprod) # Reverse order compared to per-species prob
        pcollstr[k+1] = pvecstr.copy()
        pcollcom[k+1] = pveccom.copy()

        #### 5) Rearrange communities based on new probs
        ## Standard design: emigrate lowest-scoring species, immigrate uniformly random
        remix_migrateworst(commvec, pveccom, pvecstr) # Modify commvec in-place
        commcoll[k+1] = commvec.copy()

    return( commcoll, prodcoll, pcollstr, pcollcom )


##########################################################################
##########################################################################


def repctrl_combins(df_combs,exppars):
    ########## FUNCTION TO REPEAT SIMULATION OF COMMUNITIES ##########
    ####### INPUT  #######
    #### df_combs
    #### exppars
    ####### OUTPUT #######
    ##########

    speccols = [c for c in df_combs.columns if 'Spec' in c]
    specvec  = np.unique(df_combs.loc[:,speccols])

    Nspec = len(specvec)
    Nrep  = exppars['Nrep']
    Ncomm = exppars['Ncomm']
    w_old = exppars['w_old']
    N_spc = int(np.max(df_combs.Survival)) # Assuming that at least one combination is coexistent

    pvecstr = [np.divide(1.0,Nspec),]*Nspec
    pveccom = [np.divide(1.0,Ncomm),]*Ncomm

    #### Throughout simulation, we collect per-species and community properties
    pcollstr = [pvecstr]*(Nrep+1)                 # Probability-COLLection of STRains
    pcollcom = [pveccom]*(Nrep+1)                 # Probability-COLLection of COMmunities
    commcoll = [np.zeros([Ncomm,N_spc])]*(Nrep+1) # COMMunity COLLection, ie strain identities in communities
    prodcoll = np.zeros([Nrep,Ncomm])             # PRODuction COLLection, ie per-community production

    #### 1) Allocate Ncomm communities with N_spc species per comm
    ##     randomly according to pvec
    commvec = allocatecomms(pvecstr,{'N_spc':N_spc,'Ncom':Ncomm})
    commcoll[0] = commvec.copy()
    for k in range(Nrep):
        #### 2) Simulate growth and hence community production
        commprod, survival, commidcs = simulcomms_hash(commvec, df_combs)
        prodcoll[k,:] = np.abs(commprod)

        #### 3) Find ranks with respect to community degradation
        netpen = np.divide(survival, N_spc)

        # ranks = np.argsort(np.multiply(commprod,np.max(netpen)))
        # ranks = np.argsort(np.multiply(commprod,np.min(netpen)))
        ranks  = np.argsort(np.multiply(commprod,np.power(netpen,4.0)))
        # ranks = np.argsort(np.multiply(netpen,commprod))

        #### 4) Compute new probabilities
        pvecstr = probfromranks(commvec, ranks, pvecstr, [w_old,1.0-w_old])
        pveccom = probcomm_score(commprod) # Reverse order compared to per-species prob
        pcollstr[k+1] = pvecstr.copy()
        pcollcom[k+1] = pveccom.copy()

        #### 5) Rearrange communities based on new probs
        ## Standard design: emigrate lowest-scoring species, immigrate uniformly random
        remix_ctrl(commvec,Nspec) # Modify commvec in-place
        commcoll[k+1] = commvec.copy()

    return( commcoll, prodcoll, pcollstr, pcollcom )
